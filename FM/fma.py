import openpyxl as xl
from openpyxl.styles import PatternFill, Font, colors
from openpyxl.utils import get_column_letter
from datetime import date
import shutil
import csv
import configparser
import logging
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import os
import re
from datetime import date

APP_TITLE = 'v3.2 Functionality Matrix Generator'

SHEET_1_TITLE = 'Stores_Contact-Preference_Alert'
SHEET_2_TITLE = 'Delivery_and_Payment-methods'
SHEET_3_TITLE = 'Tax_ID_number_script'
SHEET_4_TITLE = 'Gift Card Configuration'
SHEET_5_TITLE = 'Phone and Invoice Countries'
SHEET_6_TITLE = 'Payments SDK'

TITLE_SHEET_1_SECTION_1 = '### NesStore with config ###'
TITLE_SHEET_1_SECTION_2 = '### Contact Preferences Management ###'
TITLE_SHEET_1_SECTION_3 = '### DoubleOptin With GracePeriod ###'
TITLE_SHEET_1_SECTION_4 = '### Account suspension ###'
TITLE_SHEET_1_SECTION_5 = '### Contact Address Form Definition ###'
TITLE_SHEET_1_SECTION_6 = '### Contact Form Definition ###'
TITLE_SHEET_1_SECTION_7 = '### EasyOrder linked with Subscription ###'
TITLE_SHEET_1_SECTION_8 = '### Alert Management and Standing Orders ###'
TITLE_SHEET_2_SECTION_1 = '### NC2 Shipping Methods ###'
TITLE_SHEET_2_SECTION_2 = '### NC2 Payment Methods ###'
TITLE_SHEET_2_SECTION_3 = '### NC2 Payment Card Types ###'
TITLE_SHEET_2_SECTION_4 = '### NC2 3D Secure stores ###'
TITLE_SHEET_2_SECTION_5 = '### NC2 3D Secure stores ###'
TITLE_SHEET_3_SECTION_1 = '### Tax ID verification script ###'
TITLE_SHEET_3_SECTION_2 = '### State Inscription Number ###'
TITLE_SHEET_3_SECTION_3 = '### Invoicing Definition ###'
TITLE_SHEET_4_SECTION_1 = '### Gift Card Configuration ###'
TITLE_SHEET_5_SECTION_1 = '### Phone Countries ###'
TITLE_SHEET_5_SECTION_2 = '### Invoice Countries ###'
TITLE_SHEET_6_SECTION_1 = '### Payment SDK Config ###'

TITLE_ITEMS_SHEET_1_SECTION_1 = ['Store ID', 'Shipping Countries', 'Channels', 'Calleo Contact',
                                 'Calleo Contact Mobile', 'Country', 'Web Call Back', 'Live Chat', 'Machine registration', 'Store locator',
                                 'Credit Card Storage', 'Express Checkout',
                                 'Checkout Preferences Saving', 'Credit Availability', 'Remember Me',
                                 'Gift Card Purchase', 'Gift Service',
                                 'Activation Code', 'Open Shop', 'Mobile Welcome Offer', 'Machine Technologies',
                                 'EcoTax Price display',
                                 'Double Price Display', 'Fraud Management', 'Tax Display Pattern',
                                 'Split Vat Total Display',
                                 'Multi Tax Type Display', 'Total Tax Type Display', 'General Total Tax Display',
                                 'stateInscriptionNumber(uid)',
                                 'Invoicing Available', 'Invoicing Definition(PK)', 'Checkister', 'App Ranking',
                                 'Auto Login', 'Discovery Offer',
                                 'Le Club', 'Subscriptions', 'Rating Store Isolated', 'Rating Enabled For Desktop',
                                 'Rating Enabled For Mobile',
                                 'Rating Enabled For Native', 'Full Native Checkout Enabled',
                                 'Mini Cart Quotation Enabled',
                                 'Mini Cart Quotation Enabled For Anonymous', 'Passbook Notification Enabled',
                                 'Passbook Notice Confirmation Enabled',
                                 'b2b Coffee Plan Enabled', 'Guest Checkout Enabled', 'Guest Checkout Account Creation',
                                 'Passbook Enabled', 'Mini Cart Editable', 'Enable 24H Format in Delivery', 'Autodectet CC Enabled',
                                 'Technology Warning Enabled', 'Guest Checkout Map Enabled', 'Virtual Assistant Link', 
                                 'Virtual Assistant Enabled', 'B2B Subscription Config', 'Marketing Push Notification Optin',
                                 'Passbook Notification Message', 'Display App Permisions', 'Mobile Apps Font Style', 
                                 'Pup Recycling Filter Enabled','Max Large Pushes Per Technology',
                                 'Recaptcha Disable for', 'Display link for boutique checkout', 'CX Replay for app id for iOS',
                                 'CX Replay for app id for android', 'Enabled Tracking SDK', 'Enabled Adobe SDK', 'Adobe SDK ID',
                                 'ECAPI Cart Enabled', 'Fast BTG Config', 'ECAPI Cart Split Config', 'Machine Registration Step Disabled On',
                                 'Automatic Sitemap Enabled', 'contactWhatsappConfigured', 'contactwhatsappDeepLink', 'Consent Management Service Enabled','Data Protection Configuration',
                                 'Address Validation Provider', 'Address Autocomplete Provider', 'Address Map Api', 'Split Terms And Privacy On Flows']

TITLE_ITEMS_SHEET_1_SECTION_2 = ['Store ID', 'Channels', 'Enable Privacy Notice On Registration', 'Enable Registration Contact Preferences Grouped Opt-in Selection',
                                 'Enable Registration Post Mail Authorization',
                                 'Enable Registration Phone Authorization', 'Enable Registration Messaging Authorized', 'Enable Registration Use My Data For Commercial Offers Authorization',
                                 'Enable Data Profiling activities on Registration', 'Enable Registration Market Research And Satisfaction Research Authorization', 'Enable Market Research on Registration',
                                 'Enable Registration Email Authorization', 'Preselect Privacy Notice On Registration',
                                 'Preselect Registration Contact Preferences Grouped Opt-in', 'Preselect Registration Post Mail Authorization',
                                 'Preselect Registration Phone Authorization', 'Preselect Registration Messaging Authorization',
                                 'Preselect Registration Use My Data For Commercial Offers Authorization', 'Preselect Data Profiling activities on Registration',
                                 'Preselect Registration Use My Data For Market And Satisfaction Research Authorization', 'Preselect Market Research on Registration',
                                 'Preselect Registration Email Authorization', 'Enable My Account Contact Preferences Grouped Opt-in Selection',
                                 'Enable My Account Post Mail Authorization', 'Enable My Account Phone Authorization',
                                 'Enable My Account Messaging Authorization', 'Enable My Account Use My Data For Commercial Offers Authorization',
                                 'Enable Data Profiling activities on My Account', 'Enable My Account Market Research And Satisfaction Research Authorization',
                                 'Enable Market Research on My Account', 'Enable My Account Email Authorization', 'Online Invoice Configuration']
TITLE_ITEMS_SHEET_1_SECTION_3 = ['Double Optin Config code', 'double Optin Activated', 'grace Period Minutes']
TITLE_ITEMS_SHEET_1_SECTION_4 = ['Acccount suspension code', 'Account suspension Enabled']
TITLE_ITEMS_SHEET_1_SECTION_5 = ['Store ID', 'Contact Address']
TITLE_ITEMS_SHEET_1_SECTION_6 = ['Contact Form ID', 'Title', 'First Name', 'Additional Name', 'Second Name', 'email',
                                 'First Phone', 'Language']
TITLE_ITEMS_SHEET_1_SECTION_7 = ['Store ID', 'EasyOrder linked with Subscription']
TITLE_ITEMS_SHEET_1_SECTION_8 = ['Channel ID', 'Descaling Alert', 'Reorder Alert', 'Standing Orders FrontEnds',
                                 'Credit Note Info']
TITLE_ITEMS_SHEET_2_SECTION_1 = ['', 'Channels', 'NesOA ID', 'Name', 'Active', 'Authorized For Capsule',
                                 'Authorized For Machine', 'Authorized For Accessory', 'Boutique', 'Pick Up Point',
                                 'Nespresso Your Time', 'Active Recycling', 'Gift Service Compatible',
                                 'Used Preferences Saving', 'Uses Delivery Phone Number', 'Delivery Phone Number(uid)',
                                 'Delivery Phone Number(visible)', 'Delivery Phone Number(visibleInMobile)',
                                 'Delivery Phone Number(mandatory)', 'Available Payment Methods (ID:name)',
                                 'eligible For Guest Checkout']
TITLE_ITEMS_SHEET_2_SECTION_2 = ['', 'Channels', 'NesOA ID', 'Name', 'Active', 'Authorized For Capsule',
                                 'Authorized For Machine', 'Authorized For Accessory', 'Authorized For First Order',
                                 'User Authorized For Zero Amount Payment', 'Authorized For Direct Debit User',
                                 'Gift Card Compatible', 'Credit Authorized', 'Authorized For Invoice Third User',
                                 'Used Preferences Saving', 'Credit Card Authorize Only', 'Eligible For Standing Order',
                                 'is Subscription Compatible', 'enabled On Flows']
TITLE_ITEMS_SHEET_2_SECTION_3 = ['', 'nesEntity(uid)', 'name[lang=en]', 'nesOAId[unique=true]', 'installmentsEnabled',
                                 'cardFieldDefinitions(uid)']
TITLE_ITEMS_SHEET_2_SECTION_4 = ['', 'store2ThreeDS', 'triggerAmountsRuleEnabled', 'triggerAmounts',
                                 'knownSignaturesRuleEnabled', 'knownSignatureExpirationDay',
                                 'maxNumberOfKnownSignatures', 'paymentSignatureFields', 'store2CVC']
TITLE_ITEMS_SHEET_2_SECTION_5 = ['', 'Entity', 'iFrameHeight', 'iFrameWidth', 'standingOrder3DSAuthAmount']
TITLE_ITEMS_SHEET_3_SECTION_1 = ['Tax ID', 'Label', 'Regex Error Message', 'Verification Script', 'Regex']
TITLE_ITEMS_SHEET_3_SECTION_2 = ['State Inscription Number ID', 'Label', 'Regex Error Message', 'Verification Script',
                                 'Regex']
TITLE_ITEMS_SHEET_3_SECTION_3 = ['Invoicing Definition Details ID', 'Label', 'Regex Error Message',
                                 'Verification Script', 'Regex']
TITLE_ITEMS_SHEET_4_SECTION_1 = ['No Desktop Payment For (id)', 'No Mobile Payment For (id)']
TITLE_ITEMS_SHEET_5_SECTION_1 = ['', 'ID', 'nesEntityChannel', 'Name', 'isocode', 'country names', 'continental area']
TITLE_ITEMS_SHEET_5_SECTION_2 = ['ID', 'isocode', 'country name', 'continental area']

TITLE_ITEMS_SHEET_6_SECTION_1 = ['Store ID', 'Flow', 'CC Iframe Activated']

OFFSET_NESSTORE = (1, 1)

TEMPORAL_CSV_EDITED_FILE = 'temporal_edited.csv'

COLOR_SECTION = "4F81BD"
COLOR_ODD_ROWS = "0099FF"

logLevel = { "0": "NOT SET", "10": "DEBUG", "20": "INFO", "30": "WARNING", "40": "ERROR", "50": "CRITICAL"}
logger = logging.getLogger(__name__)
logging.basicConfig(format='%(asctime)s:%(name)s:%(message)s', level=logging.DEBUG)
logger.info(logLevel)
logger.info(f"Log set to: {logger.root.level}")
logger.error("*** Let's generate the functionality matrix from an CSV Export file ***")


class FunctionalityMatrix:
    def __init__(self, csv_file, excel_out_file):
        # Save csv file name
        self.csv_file = csv_file
        # Create temporal csv_editd_filñe
        self.csv_file_edited = csv_file
        # Save excel output file name
        self.excel_out_file = excel_out_file
        # Create a workbook (excel type)
        self.wb = xl.Workbook()
        # Create tab sheets and rename them
        ss = self.wb.active   # wb['Sheet']
        ss.title = SHEET_1_TITLE
        self.wb.create_sheet(SHEET_2_TITLE)
        self.wb.create_sheet(SHEET_3_TITLE)
        self.wb.create_sheet(SHEET_4_TITLE)
        self.wb.create_sheet(SHEET_5_TITLE)
        self.wb.create_sheet(SHEET_6_TITLE)
        logger.info('Creating workbook with sheets named:')
        logger.info(self.wb.sheetnames)

        self.util_remove_unused_entity_channels_and_add_space_before_insert_update()

        # self.csv_file_in = open(self.csv_file_edited, 'r', newline='', encoding='utf-8')
        # self.csv_reader = csv.reader(self.csv_file_in, delimiter=';', quoting=csv.QUOTE_NONE)

    def __del__(self):
        try:
            # Extract the folder to save the excel PATH
            path, name = os.path.split(self.csv_file)
            # Assigne to NAME the name of the excel file with the current date in dd-mm-yyyy format.
            name = date.today().strftime('%d-%m-%Y') +'_NC2_Functionality_Matrix.xlsx'
            # join the path and name
            excel_file_name = os.path.join(path, name)

            # format column width and zoom
            for ws in self.wb.worksheets:
                ws.sheet_view.zoomScale = 85
                for i, column_number in enumerate(ws.columns):
                    ws.column_dimensions[get_column_letter(i+1)].width = 25
                ws.column_dimensions['A'].width = 65

                # for i in range(2, 200):
                #     ws.column_dimensions[get_column_letter].width = 25

            # Save the excel file
            logger.info('Writing EXCEL file...')
            self.wb.save(excel_file_name)
            logger.info('Saved!')
            # Close file not needed
            # self.csv_file_in.close()
            # Removing temporal file
            # uncomment after test
            os.remove(self.csv_file_edited)
            # open the destination folder
            logger.info('Open folder where EXCEL FILE is...')
            #os.startfile(path)
            logger.info(f'Path: {path}')
            logger.info(f'{name}')
            logger.info('Excel File created sucessfully!')
            return 1
        except:
            logger.error(f'Error saving EXCEL file. Probably file is open. Close and Run again. EXCEPTION: {ValueError}')
            return -1




    def util_remove_unused_entity_channels_and_add_space_before_insert_update(self):
        #fn2 = "out.csv" was temporal, now using self.csv_file_edited

        # with open(self.csv_file, encoding='utf8') as fn, open(fn2, 'w', encoding='utf8') as fn2:
        #     clean = fn.read().replace('\ninsert', '\n;\ninsert')
        #     fn2.write(clean)
        self.csv_file_edited = TEMPORAL_CSV_EDITED_FILE
        with open(self.csv_file, encoding='utf8') as fn, open(self.csv_file_edited, 'w', encoding='utf8') as fn2:
            for line in fn:
                line = re.sub(r'^.*EntityChannel_INT_B2B.*$\n', '', line)
                line = re.sub(r'^.*EntityChannel_INT_B2C.*$\n', '', line)
                line = re.sub(r'insert', r';\ninsert', line)
                fn2.writelines(line)

    def util_format_line(self, sheet_name, position, line, color_background, color_font):
        # Point the sheet to work with
        sheet = self.wb[sheet_name]

        # removing blanks from the begining of the list
        for i, value in enumerate(line):
            if value == '':
                line.pop(i)
                break
        # Blue FM color r0g153b255 0x0099FF
        for i, value in enumerate(line):
            sheet.cell(position[0], position[1] + i).fill = PatternFill(start_color=color_background,
                                                                        end_color=color_background,
                                                                        fill_type='solid')
            sheet.cell(position[0], position[1] + i).font = Font(color=color_font)
    pass

    def util_format_title_section(self, sheet_name, section_name, section_position):
        # Point the sheet to work with
        sheet = self.wb[sheet_name]

        sheet.cell(section_position[0], section_position[1]).value = section_name


        # Blue FM color r79g129b189 0x4F81BD
        for i in range(100):
            sheet.cell(section_position[0], section_position[1]+i).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type='solid')
            sheet.cell(section_position[0], (section_position[1])+i).font = Font(color=colors.WHITE)       
        


    def util_line_write_from_csv_to_excel_not_transpose(self, sheet_name, offset, offsetElementNum, firstElement, line):
        # Point the sheet to work with
        sheet = self.wb[sheet_name]

        j=0
        for i in range(firstElement, len(line)):
            sheet.cell(offset[0] + offsetElementNum, offset[1] + j).value = line[i]
            j += 1


    def util_line_write_from_csv_to_excel_transpose(self, sheet_name, offset, offsetElementNum, firstElement, line):
        # Point the sheet to work with
        sheet = self.wb[sheet_name]

        j=0
        for i in range(firstElement, len(line)):
            sheet.cell(offset[0] + j, offset[1] + offsetElementNum).value = line[i]
            j += 1

    # Creates the filters in every colum of a block, with the start cell and end cell of the block
    def util_filter_columns(self, sheetName, cellBlockStart, cellBlockEnd):
        letterStart = get_column_letter(cellBlockStart[0])
        letterEnd   = get_column_letter(cellBlockEnd[0])
        numberStart = str(cellBlockStart[1])
        numberEnd   = str(cellBlockEnd[1])
        logger.info(f'Autofilter block will start in this cell CellBlockStart: {cellBlockStart}')
        logger.info(f'Autofilter block will end in this cell CellBlockEnd:     {cellBlockEnd}')
        self.wb[sheetName].auto_filter.ref = letterStart + numberStart + ':' + letterEnd + numberEnd
        logger.info(f'Column block with letters: {letterStart}{numberStart}:{letterEnd}{numberEnd}')


    def sheet_1_section_1_nesStore(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update nesstore' in line[0].lower() and 'shippingcountries(isocode)' in line[2].lower():
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_1, (1, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update NesStore')
                        # This is the title items of the sheet 1 section 1
                        line = TITLE_ITEMS_SHEET_1_SECTION_1
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (2, 1), NesStoreElementNum, 0, line)
                        NesStoreElementNum+=1
                        line = next(self.csv_reader)
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (2, 1), NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 1 Section 1 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 1 EXCEPTION: {ValueError}')
            return -1

    def sheet_1_section_2_nesStore(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update nesstore' in line[0].lower() and 'nesentitychannel(id)' in line[2].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_2, (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update NesStore')
                        # This is for v1.2 and older
                        # #line = ['Store ID', 'Channels', 'Contact Channel Selection', 'Enable Registration Post Mail Authorization', 'Enable Registration Phone Authorization', 'Enable Registration Messaging Authorized', 'Data For Marketing Authorization', 'Commercial Offers', 'Enable Registration Market Research And Satisfaction Research Authorization', 'Terms And Conditions Acceptance', 'Enable Registration Email Authorization', 'Preselect Contact Channel Selection', 'Preselect Post Mail Authorization', 'Preselect Registration Phone Authorization', 'Preselect Registration Messaging Authorization', 'Preselect Data For Marketing Authorization', 'Preselect Registration Use My Data For Commercial Offers Authorization', 'Preselect Registration Use My Data For Market And Satisfaction Research Authorization', 'Preselect Terms And Conditions Acceptance', 'Preselect Email Authorization', 'My Account Contact Channel Selection', 'My Account Post Mail Authorization', 'My Account Phone Authorization', 'My Account Messaging Authorization', 'My Account Use Data For Marketing Authorization', 'My Account Commercial Offers', 'My Account Enable Satisfaction Research', 'My Account Email Authorization', 'My Account Magazine Subscription']
                        # This is for v1.3
                        line = TITLE_ITEMS_SHEET_1_SECTION_2
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1), NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1), NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 1 Section 2 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 2 EXCEPTION: {ValueError}error')
            return -1
    def sheet_1_section_3_DoubleOptinConfig(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader=csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                for line in self.csv_reader:
                    if 'insert_update DoubleOptinConfig'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_3, (self.row_number_start, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update DoubleOptinConfig')
                        line = TITLE_ITEMS_SHEET_1_SECTION_3
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),NesStoreElementNum, 1, line)
                            NesStoreElementNum +=1
                            try:
                                line = next(self.csv_reader)
                            except:
                                break
                        logger.info(f'Sheet 1 Section 3 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1    
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 3 EXCEPTION: {ValueError}')
            return -1
    def sheet_1_section_8_NesEntityChannel(self):
        # insert_update NesEntityChannel
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update nesentitychannel' in line[0].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_8, (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update NesEntityChannel')
                        line = TITLE_ITEMS_SHEET_1_SECTION_8
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                         NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '' and line[1] != '':

                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                             NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 1 Section 8 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 8 EXCEPTION: {ValueError}')
            return -1

        pass

    def sheet_1_section_4_accountSuspension(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader=csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                for line in self.csv_reader:
                    if 'insert_update AccountSuspension'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_4, (self.row_number_start, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update accountSuspension')
                        line = TITLE_ITEMS_SHEET_1_SECTION_4
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),NesStoreElementNum, 1, line)
                            NesStoreElementNum +=1
                            try:
                                line = next(self.csv_reader)
                            except:
                                break
                        logger.info(f'Sheet 1 Section 4 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1    
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 4 EXCEPTION: {ValueError}')
            return -1

    def sheet_1_section_7_Subscription(self):
        # insert_update NesEntityChannel
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update subscriptionrecurringorderconfig' in line[0].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_7,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update subscriptionrecurringorderconfig')
                        line = TITLE_ITEMS_SHEET_1_SECTION_7
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                         NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                             NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 1 Section 7 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 7 EXCEPTION: {ValueError}')
            return -1

        pass

    def sheet_1_section_5_contactAddressFormDefinition(self):
        # insert_update NesEntityChannel
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update NesStore'.lower() in line[0].lower() and 'contactAddressFormDefinition(uid)'.lower() in line[2].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_5,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update subscriptionrecurringorderconfig')
                        line = TITLE_ITEMS_SHEET_1_SECTION_5
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                         NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                             NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 1 Section 5 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 5 EXCEPTION: {ValueError}')
            return -1

        pass

    def sheet_1_section_6_contactFormDefinition(self):
        # insert_update NesEntityChannel
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update ContactFormDefinition'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_1_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_1_TITLE, TITLE_SHEET_1_SECTION_6,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update contactFormDefinition')
                        line = TITLE_ITEMS_SHEET_1_SECTION_6
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                         NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                       
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_1_TITLE, (self.row_number_start + 1, 1),
                                                                             NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            try:
                                line = next(self.csv_reader)
                            except:
                                # Cannot control End Of File, so I handle the exception.
                                break                        
                        
                        logger.info(f'Sheet 1 Section 6 Row last index={self.wb[SHEET_1_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 1 Section 6 EXCEPTION: {ValueError}')
            return -1

        pass

    

    

    def sheet_2_section_1_ShippingMethod(self):

        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update shippingmethod' in line[0].lower():
                        self.util_format_title_section(SHEET_2_TITLE, TITLE_SHEET_2_SECTION_1, (1, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update ShippingMethod')
                        line = TITLE_ITEMS_SHEET_2_SECTION_1
                        numberOfColumns = len(line)
                        self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE, (1, 2), NesStoreElementNum, 1, line)
                        self.util_format_line(SHEET_2_TITLE, (1, 2), line, COLOR_SECTION, colors.WHITE)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE, (1, 2), NesStoreElementNum, 1, line)
                            # The following lines are for formating in blue or white depending if odd or even row.
                            # For future releases.
                            # try:
                            #     if NesStoreElementNum %2 == 1:
                            #         self.util_format_line(SHEET_2_TITLE, (2 + NesStoreElementNum, 2), line, COLOR_ODD_ROWS, colors.BLACK)
                            # except ValueError:
                            #     logger.error(f"Exception %2 code: {ValueError}")

                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 2 Section 1 Row last index={self.wb[SHEET_2_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')

                        # Filter colums from first section
                        # We are commenting this. Since the Openxl library don't support more than one in the same sheet
                        # It is better dont have any, and do it manually afterwards.
                        # self.util_filter_columns(SHEET_2_TITLE, (2, 2), (numberOfColumns + 2, self.wb[SHEET_2_TITLE].max_row) )


            return 1
        except ValueError:
            logger.error(f'"Error Sheet 2 Section 1 EXCEPTION: {ValueError}')
            return -1



        pass

    def sheet_2_section_2_PatymentMethod(self):
        # insert_update NesEntityChannel
        logger.info(f'Begin step 1 CSV file number: {self.csv_reader.line_num}')
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update paymentmethod'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_2_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_2_TITLE, TITLE_SHEET_2_SECTION_2,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update PaymentMethod')
                        line = TITLE_ITEMS_SHEET_2_SECTION_2
                        numberOfColumns = len(line)
                        rowStart = self.row_number_start + 1
                        self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE,
                                                                             (self.row_number_start + 0, 2),
                                                                             NesStoreElementNum, 1, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[1] != '':
                            self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE,
                                                                                 (self.row_number_start + 0, 2),
                                                                                 NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 2 Section 2 Row last index={self.wb[SHEET_2_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')

                        # Filter colums from first section
                        # We are commenting this. Since the Openxl library don't support more than one in the same sheet
                        # It is better dont have any, and do it manually afterwards.
                        #self.util_filter_columns(SHEET_2_TITLE, (2, rowStart), (numberOfColumns + 2, self.wb[SHEET_2_TITLE].max_row))

            return 1
        except ValueError:
            logger.error(f'"Error Sheet 2 Section 2 EXCEPTION: {ValueError}')
            return -1

        pass

    def sheet_2_section_3_PaymentCardTypes(self):
        # insert_update NesEntityChannel
        logger.info(f'Begin step 3 CSV file number: {self.csv_reader.line_num}')
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update paymentcardtype'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_2_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_2_TITLE, TITLE_SHEET_2_SECTION_3,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update PaymentMethod')
                        line = TITLE_ITEMS_SHEET_2_SECTION_3
                        self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE, (self.row_number_start + 0, 2),
                                                                             NesStoreElementNum, 1, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '':
                            self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE, (self.row_number_start + 0, 2),
                                                                                 NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 2 Section 3 Row last index={self.wb[SHEET_2_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')

            return 1
        except ValueError:
            logger.error(f'"Error Sheet 2 Section 3 EXCEPTION: {ValueError}')
            return -1

        pass

    def sheet_2_section_4_PaymentSignatureConfiguration(self):
        # insert_update NesEntityChannel
        logger.info(f'Begin step 3 CSV file number: {self.csv_reader.line_num}')
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update paymentsignatureconfiguration'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_2_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_2_TITLE, TITLE_SHEET_2_SECTION_4,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update paymentSignatureConfiguration')
                        line = TITLE_ITEMS_SHEET_2_SECTION_4
                        self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE, (self.row_number_start + 0, 2),
                                                                             NesStoreElementNum, 1, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '':
                            self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE,
                                                                                 (self.row_number_start + 0, 2),
                                                                                 NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 2 Section 4 Row last index={self.wb[SHEET_2_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')

            return 1
        except ValueError:
            logger.error(f'"Error Sheet 2 Section 4 EXCEPTION: {ValueError}')
            return -1

    pass

    def sheet_2_section_5_HOPConfiguration(self):
        # insert_update HOPConfiguration
        logger.info(f'Begin tab 2 section 5 CSV file number: {self.csv_reader.line_num}')
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                   if 'insert_update HOPConfiguration'.lower() in line[0].lower():
                        self.row_number_start = self.wb[SHEET_2_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_2_TITLE, TITLE_SHEET_2_SECTION_5,
                                                       (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update HOPConfiguration')
                        line = TITLE_ITEMS_SHEET_2_SECTION_5
                        self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE, (self.row_number_start + 0, 2),
                                                                             NesStoreElementNum, 1, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '':
                            self.util_line_write_from_csv_to_excel_not_transpose(SHEET_2_TITLE,
                                                                                 (self.row_number_start + 0, 2),
                                                                                 NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 2 Section 5 Row last index={self.wb[SHEET_2_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')

            return 1
        except ValueError:
            logger.error(f'"Error Sheet 2 Section 5 EXCEPTION: {ValueError}')
            return -1

        pass

    # This function look for the FIRST insert_update FormatAttribute.
    # There is no way to distinct between the three of them. Just the order.
    def sheet_3_section_1_FormAttribute(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                # S'HA D'OBRIR EL CSV AMB " COM A INDICADOR DE TEXTE, PERQUE ACCEPTI ELS ; COM A TEXTE

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update formattribute'.lower() in line[0].lower():
                        self.util_format_title_section(SHEET_3_TITLE, TITLE_SHEET_3_SECTION_1, (1, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update FormAttribute')
                        line = TITLE_ITEMS_SHEET_3_SECTION_1
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_3_TITLE, (2, 1), NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_3_TITLE, (2, 1), NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)

                        logger.info(f'Sheet 3 Section 1 Row last index={self.wb[SHEET_3_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')

                        # We sont want iteration 2 and 3 to be done. Thats the reason BREAK
                        break
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 3 Section 1 EXCEPTION: {ValueError}')
            return -1
        pass

    # This function look for the SECOND insert_update FormatAttribute.
    # There is no way to distinct between the three of them. Just the order.
    def sheet_3_section_2_FormAttribute(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                # S'HA D'OBRIR EL CSV AMB " COM A INDICADOR DE TEXTE, PERQUE ACCEPTI ELS ; COM A TEXTE
                number_of_form_attribute = 0
                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update formattribute'.lower() in line[0].lower():
                        if number_of_form_attribute < 1:
                            number_of_form_attribute += 1
                        else:
                            self.row_number_start = self.wb[SHEET_3_TITLE].max_row + 2
                            self.util_format_title_section(SHEET_3_TITLE, TITLE_SHEET_3_SECTION_2, (self.row_number_start, 1))
                            NesStoreElementNum = 0
                            logger.debug('insert_update FormAttribute')
                            line = TITLE_ITEMS_SHEET_3_SECTION_2
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_3_TITLE, (self.row_number_start + 1, 1), NesStoreElementNum, 0,
                                                                             line)
                            NesStoreElementNum += 1
                            line = next(self.csv_reader)
                            while (line[0] == '' or 'StateInscriptionFormAttribute' in line[0]):
                                self.util_line_write_from_csv_to_excel_transpose(SHEET_3_TITLE,
                                                                                 (self.row_number_start + 1, 1),
                                                                                 NesStoreElementNum, 1,
                                                                                 line)
                                NesStoreElementNum += 1

                                line = next(self.csv_reader)

                            logger.info(f'Sheet 3 Section 1 Row last index={self.wb[SHEET_3_TITLE].max_row}')
                            logger.info(f'CSV file number: {self.csv_reader.line_num}')

                            # We sont want iteration 3 to be done. Thats the reason BREAK
                            break

            return 1
        except ValueError:
            logger.error(f'"Error Sheet 3 Section 1 EXCEPTION: {ValueError}')
            return -1
        pass

    # This function look for the FIRST insert_update FormatAttribute.
    # There is no way to distinct between the three of them. Just the order.
    def sheet_3_section_3_FormAttribute(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                # S'HA D'OBRIR EL CSV AMB " COM A INDICADOR DE TEXTE, PERQUE ACCEPTI ELS ; COM A TEXTE
                number_of_form_attribute = 0
                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update formattribute'.lower() in line[0].lower():
                        if number_of_form_attribute < 2:
                            number_of_form_attribute += 1
                        else:
                            self.row_number_start = self.wb[SHEET_3_TITLE].max_row + 2
                            self.util_format_title_section(SHEET_3_TITLE, TITLE_SHEET_3_SECTION_3, (self.row_number_start, 1))
                            NesStoreElementNum = 0
                            logger.debug('insert_update FormAttribute')
                            line = TITLE_ITEMS_SHEET_3_SECTION_3
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_3_TITLE, (self.row_number_start + 1, 1), NesStoreElementNum, 0,
                                                                             line)
                            NesStoreElementNum += 1
                            line = next(self.csv_reader)
                            while line[0] == '' and line[1] != '':
                                self.util_line_write_from_csv_to_excel_transpose(SHEET_3_TITLE, (self.row_number_start + 1, 1),
                                                                                 NesStoreElementNum, 1, line)
                                NesStoreElementNum += 1

                                line = next(self.csv_reader)

                            logger.info(f'Sheet 3 Section 1 Row last index={self.wb[SHEET_3_TITLE].max_row}')
                            logger.info(f'CSV file number: {self.csv_reader.line_num}')

                            # WE just break. We dont need more iterations.
                            break

            return 1
        except ValueError:
            logger.error(f'"Error Sheet 3 Section 1 EXCEPTION: {ValueError}')
            return -1
        pass

    def sheet_4_section_1_GiftCardConfiguration(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update giftcardconfiguration'.lower() in line[0].lower():
                        self.util_format_title_section(SHEET_4_TITLE, TITLE_SHEET_4_SECTION_1, (1, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update giftCardConfiguration')
                        line = TITLE_ITEMS_SHEET_4_SECTION_1
                        self.util_line_write_from_csv_to_excel_not_transpose(SHEET_4_TITLE, (2, 1), NesStoreElementNum, 0, line)
                        # for splited text NesStoreElement must be 0, not 1 for first element
                        # because is not using the line from de csv, it is using the text_splited, and the [0] is data too.
                        # NesStoreElementNum+=1>
                        line = next(self.csv_reader)
                        # No Desktop Payment For (id) block
                        text_splited = line[2].split(',')
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_4_TITLE, (3, 1), NesStoreElementNum, 0, text_splited)
                        NesStoreElementNum += 1

                        text_splited = line[3].split(',')
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_4_TITLE, (3, 1), NesStoreElementNum, 0,
                                                                         text_splited)

                        logger.info(f'Sheet 4 Section 1 Row last index={self.wb[SHEET_4_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 4 Section 1 EXCEPTION: {ValueError}')
            return -1
        pass

    def sheet_5_section_1_PhoneCountries(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update NesStore'.lower() in line[0].lower() and 'phoneCountries(isocode)'.lower() in line[4].lower():
                        self.util_format_title_section(SHEET_5_TITLE, TITLE_SHEET_5_SECTION_1, (1, 1))
                        NesStoreElementNum = 0
                        logger.debug('insert_update phone countries')

                        line = TITLE_ITEMS_SHEET_5_SECTION_1
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_5_TITLE, (2, 1), NesStoreElementNum, 1, line)

                        line = next(self.csv_reader)
                        NesStoreElementNum += 1
                        while line[0] == '' and line[1] != '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_5_TITLE, (2, 1), NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)
                        logger.info(f'Sheet 5 Section 1 Row last index={self.wb[SHEET_5_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 5 Section 1 EXCEPTION: {ValueError}')
            return -1
    def sheet_5_section_2_InvoiceCountries(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update nesstore'.lower() in line[0].lower() and 'invoiceCountries(isocode)'.lower() in line[2].lower():
                        self.row_number_start = self.wb[SHEET_5_TITLE].max_row + 2
                        self.util_format_title_section(SHEET_5_TITLE, TITLE_SHEET_5_SECTION_2, (self.row_number_start, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update NesStore')
                        line = TITLE_ITEMS_SHEET_5_SECTION_2
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_5_TITLE, (self.row_number_start + 1, 1), NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_5_TITLE, (self.row_number_start + 1, 1), NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)
                            # This is a very bad programing thing.... but i dont have time
                            # When it found the first element to be aborted, it aborts.
                            if 'NesStore_ae'.lower() in line[1].lower():
                                break

                        logger.info(f'Sheet 5 Section 2 Row last index={self.wb[SHEET_5_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 5 Section 2 EXCEPTION: {ValueError}error')
            return -1
    
    def sheet_6_section_1_PaymentSDK(self):
        try:
            with open(self.csv_file_edited, encoding='utf8') as file_name:
                self.csv_reader = csv.reader(file_name, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                for line in self.csv_reader:
                    # I 'lower()' the string, so any change in the future of capital letters doesn't affect.
                    if 'insert_update PaymentSDKConfig'.lower() in line[0].lower() :
                        self.util_format_title_section(SHEET_6_TITLE, TITLE_SHEET_6_SECTION_1, (1, 1))
                        NesStoreElementNum = 0

                        logger.debug('insert_update PaymentSDKConfig')
                        line = TITLE_ITEMS_SHEET_6_SECTION_1
                        self.util_line_write_from_csv_to_excel_transpose(SHEET_6_TITLE, (2, 1), NesStoreElementNum, 0, line)
                        NesStoreElementNum += 1
                        line = next(self.csv_reader)

                        while line[0] == '':
                            self.util_line_write_from_csv_to_excel_transpose(SHEET_6_TITLE, (2, 1), NesStoreElementNum, 1, line)
                            NesStoreElementNum += 1

                            line = next(self.csv_reader)
                            # This is a very bad programing thing.... but i dont have time
                            # When it found the first element to be aborted, it aborts.
                            if 'NesStore_ae'.lower() in line[1].lower():
                                break

                        logger.info(f'Sheet 6 Section 1 Row last index={self.wb[SHEET_6_TITLE].max_row}')
                        logger.info(f'CSV file number: {self.csv_reader.line_num}')
                        return 1
            return 1
        except ValueError:
            logger.error(f'"Error Sheet 6 Section 1 EXCEPTION: {ValueError}error')
            return -1
    

    def step_test(self):
        pass

def click_run():
    # Create Functionally Matrix Object
    fm = FunctionalityMatrix(filename_csv, '')
    # FunctionalityMatrix sections. This could be done inside FM class, and make runAll() function
    # SHEET 1
    fm.sheet_1_section_1_nesStore()
    fm.sheet_1_section_2_nesStore()
    fm.sheet_1_section_3_DoubleOptinConfig()
    fm.sheet_1_section_4_accountSuspension()
    fm.sheet_1_section_5_contactAddressFormDefinition()
    fm.sheet_1_section_6_contactFormDefinition()
    fm.sheet_1_section_7_Subscription()
    fm.sheet_1_section_8_NesEntityChannel()
    # SHEET 2
    fm.sheet_2_section_1_ShippingMethod()
    fm.sheet_2_section_2_PatymentMethod() 
    fm.sheet_2_section_3_PaymentCardTypes()
    fm.sheet_2_section_4_PaymentSignatureConfiguration()
    fm.sheet_2_section_5_HOPConfiguration()
    # SHEET 3
    fm.sheet_3_section_1_FormAttribute()  # TAX ID verification script
    fm.sheet_3_section_2_FormAttribute()
    fm.sheet_3_section_3_FormAttribute()

    # SHEET 4
    fm.sheet_4_section_1_GiftCardConfiguration()

    # SHEET 5
    fm.sheet_5_section_1_PhoneCountries()
    fm.sheet_5_section_2_InvoiceCountries()

    # SHEET 6
    fm.sheet_6_section_1_PaymentSDK()


    # Testing
    fm.step_test()

    # Free object
    del fm



def check_run_avaliable():
    try:
        condition = bool(filename_csv)  # and filename_fm_excel)  # strings can be checked as bool if they are empty or not.
        if condition:
            run_button.config(state='normal')
        else:
            run_button.config(state='disabled')
    except:
        logger.info('Error Checking run avaliable')

def click_data_csv():
    global filename_csv

    filename_csv = filedialog.askopenfilename(initialdir=os.getcwd(), title='Choose your exported data csv file',
                                                    filetypes=(("csv files", "*.csv"), ("all files", "*.*")))
    path, filename = os.path.split(filename_csv)

    input_csv_label.config(text='          ' + filename + '          ')

    check_run_avaliable()
    logger.info(f'CSV File is: {filename_csv}')


def init_GUI():
    global run_button
    global input_csv_label
    global input_excel_label
    global run_label

    root = tk.Tk()
    root.resizable(0, 0)
    root.title(APP_TITLE)
    root.geometry('+%d+%d' % (675, 250))

    input_csv_button = tk.Button(root,  anchor='w', text='Select Exported data', command=click_data_csv)
    input_csv_button.grid(row=1, column=0)
    input_csv_label = tk.Label(root, text='          <csv filename>          ')
    input_csv_label.grid(row=1, column=1)

    # input_excel_button = tk.Button(root, anchor='w', text='Select last F. Matrix.xlsx', command=click_fm_excel)
    # input_excel_button.grid(row=2, column=0)
    # input_excel_label = tk.Label(root, padx=20, text='<excel name>')
    # input_excel_label.grid(row=2, column=1)

   #  edit_name_out = tk.Entry(root)
   #  edit_name_out.grid(row=3, column=0)
   # # edit_name_out.delete(0, END)
   #  today = f'{date.today()}_NC2_Functionality_Matrix.xlsx'
   #  edit_name_out.insert(0, today)

    run_button = tk.Button(root, padx=30, anchor='w', text='Run...',  command=click_run, state=tk.DISABLED)
    run_button.grid(row=6, column=0)
    run_label = tk.Label(root, text='')
    run_label.grid(row=6, column=1)

    root.mainloop()


if __name__ == "__main__":
    init_GUI()

